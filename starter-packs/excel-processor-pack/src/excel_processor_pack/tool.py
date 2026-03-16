"""Create, read, and append data to Excel workbooks."""

from __future__ import annotations

import os
import tempfile

from openpyxl import Workbook, load_workbook


def _read_workbook(file_path: str, sheet_name: str) -> dict:
    """Read data from an existing Excel workbook."""
    file_path = os.path.abspath(file_path)
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames

    # Pick the requested sheet, or the first one
    if sheet_name in sheets:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    data: list[list] = []
    for row in ws.iter_rows(values_only=True):
        data.append([cell if cell is not None else "" for cell in row])

    wb.close()

    return {
        "data": data,
        "sheets": sheets,
        "rows": len(data),
        "columns": max((len(r) for r in data), default=0),
    }


def _create_workbook(
    data: list[list] | None,
    output_path: str,
    sheet_name: str,
) -> dict:
    """Create a new Excel workbook with the given data."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    rows_written = 0
    max_cols = 0

    if data:
        for row in data:
            ws.append(row)
            rows_written += 1
            max_cols = max(max_cols, len(row))

    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()

    return {
        "output_path": output_path,
        "rows": rows_written,
        "columns": max_cols,
    }


def _append_to_workbook(
    file_path: str,
    data: list[list] | None,
    output_path: str,
    sheet_name: str,
) -> dict:
    """Append rows to an existing workbook."""
    file_path = os.path.abspath(file_path)
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    rows_added = 0
    max_cols = 0

    if data:
        for row in data:
            ws.append(row)
            rows_added += 1
            max_cols = max(max_cols, len(row))

    if not output_path:
        output_path = file_path

    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()

    return {
        "output_path": output_path,
        "rows": ws.max_row,
        "columns": ws.max_column,
    }


def run(
    operation: str = "create",
    file_path: str = "",
    data: list[list] | None = None,
    output_path: str = "",
    sheet_name: str = "Sheet1",
) -> dict:
    """Create, read, or append data to an Excel workbook.

    Args:
        operation: One of "create", "read", "append".
        file_path: Path to an existing Excel file (for "read" and "append").
        data: 2-D list of cell values (for "create" and "append").
        output_path: Where to save the file. Auto-generated for "create" if empty.
                     For "append", defaults to overwriting file_path.
        sheet_name: Worksheet name to target.

    Returns:
        For "read": {"data": list, "sheets": list, "rows": int, "columns": int}
        For "create"/"append": {"output_path": str, "rows": int, "columns": int}
    """
    operation = operation.strip().lower()

    if operation == "read":
        if not file_path:
            raise ValueError("file_path is required for the 'read' operation.")
        return _read_workbook(file_path, sheet_name)

    if operation == "create":
        if not output_path:
            output_path = os.path.join(tempfile.gettempdir(), "workbook.xlsx")
        return _create_workbook(data, output_path, sheet_name)

    if operation == "append":
        if not file_path:
            raise ValueError("file_path is required for the 'append' operation.")
        return _append_to_workbook(file_path, data, output_path, sheet_name)

    raise ValueError(
        f"Unknown operation '{operation}'. Choose from: create, read, append."
    )
